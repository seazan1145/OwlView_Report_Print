from __future__ import annotations

import base64
import os
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from PIL import Image

from .models import CommonConfig, PartConfig

try:
    import win32print  # type: ignore
except ImportError:  # pragma: no cover
    win32print = None


@dataclass
class ExternalTools:
    curl: Path
    sumatra: Path


@dataclass
class CurlResult:
    command_summary: str
    returncode: int
    stdout: str
    stderr: str


def resolve_tool_path(configured: str, fallback: Path, exe_name: str) -> Path:
    if configured.strip():
        return Path(configured).expanduser()
    if fallback.exists():
        return fallback
    from shutil import which

    candidates: list[str] = [exe_name]
    if exe_name.lower().endswith(".exe"):
        candidates.append(exe_name[:-4])
    for name in candidates:
        found = which(name)
        if found:
            return Path(found)
    return fallback


def printer_list() -> list[str]:
    if not win32print:
        return []
    return sorted({p[2] for p in win32print.EnumPrinters(2)})


def print_with_sumatra(sumatra: Path, pdf_path: Path, printer: str, copies: int) -> None:
    creationflags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
    subprocess.run(
        [str(sumatra), "-silent", "-exit-on-print", "-print-to", printer, "-print-settings", f"{copies}x", str(pdf_path)],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        creationflags=creationflags,
    )


def save_pdf(driver, dest: Path, part: PartConfig) -> None:
    prefs = {
        "paperWidth": part.paper_width,
        "paperHeight": part.paper_height,
        "marginTop": part.margin_top,
        "marginBottom": part.margin_bottom,
        "marginLeft": part.margin_left,
        "marginRight": part.margin_right,
        "printBackground": True,
        "scale": part.scale / 100.0,
        "landscape": part.orientation == "landscape",
    }
    if part.print_range.strip():
        prefs["pageRanges"] = part.print_range.strip()
    data = driver.execute_cdp_cmd("Page.printToPDF", prefs)["data"]
    dest.write_bytes(base64.b64decode(data))


def save_jpg_from_screenshot(driver, dest: Path, quality: int = 90) -> None:
    png = driver.get_screenshot_as_png()
    tmp = dest.with_suffix(".tmp.png")
    tmp.write_bytes(png)
    with Image.open(tmp) as im:
        rgb = im.convert("RGB")
        rgb.save(dest, "JPEG", quality=quality)
    tmp.unlink(missing_ok=True)


def save_jpg_from_pdf(driver, pdf_path: Path, dest: Path, quality: int = 90) -> None:
    convert_pdf_first_page_to_jpg(pdf_path, dest, quality=quality)


def _convert_pdf_with_pymupdf(pdf_path: Path, dest: Path, quality: int, dpi: int) -> bool:
    try:
        import fitz  # type: ignore
    except ImportError:
        return False
    doc = fitz.open(pdf_path)
    try:
        if doc.page_count < 1:
            raise RuntimeError(f"PDFページがありません: {pdf_path}")
        page = doc.load_page(0)
        scale = max(1.0, dpi / 72.0)
        mat = fitz.Matrix(scale, scale)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        mode = "RGB" if pix.n < 4 else "RGBA"
        img = Image.frombytes(mode, [pix.width, pix.height], pix.samples)
        img.convert("RGB").save(dest, "JPEG", quality=quality)
        return True
    finally:
        doc.close()


def _convert_pdf_with_pdf2image(pdf_path: Path, dest: Path, quality: int, dpi: int) -> bool:
    try:
        from pdf2image import convert_from_path  # type: ignore
    except ImportError:
        return False
    with tempfile.TemporaryDirectory() as td:
        images = convert_from_path(
            str(pdf_path),
            dpi=dpi,
            first_page=1,
            last_page=1,
            fmt="png",
            output_folder=td,
        )
        if not images:
            raise RuntimeError(f"PDFの画像化に失敗しました: {pdf_path}")
        images[0].convert("RGB").save(dest, "JPEG", quality=quality)
    return True


def convert_pdf_first_page_to_jpg(pdf_path: Path, dest: Path, quality: int = 90, dpi: int = 200) -> None:
    """
    PDFの1ページ目をJPG化する。現仕様では1ページ目のみ対応。
    """
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDFが見つかりません: {pdf_path}")
    if _convert_pdf_with_pymupdf(pdf_path, dest, quality, dpi):
        return
    if _convert_pdf_with_pdf2image(pdf_path, dest, quality, dpi):
        return
    raise RuntimeError("PDF->JPG変換ライブラリがありません。PyMuPDF(fitz) もしくは pdf2image(+poppler) を導入してください。")


def render_pdf_first_page_image(pdf_path: Path, dpi: int = 144) -> Image.Image:
    try:
        import fitz  # type: ignore
    except ImportError as exc:
        raise RuntimeError("PDFプレビューにはPyMuPDF(fitz)が必要です。") from exc
    doc = fitz.open(pdf_path)
    try:
        if doc.page_count < 1:
            raise RuntimeError(f"PDFページがありません: {pdf_path}")
        page = doc.load_page(0)
        scale = max(1.0, dpi / 72.0)
        mat = fitz.Matrix(scale, scale)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        mode = "RGB" if pix.n < 4 else "RGBA"
        return Image.frombytes(mode, [pix.width, pix.height], pix.samples).convert("RGB")
    finally:
        doc.close()


def validate_ftp_path_template(path_template: str) -> list[str]:
    issues: list[str] = []
    v = path_template.strip()
    if not v:
        issues.append("FTP Path が未設定です")
        return issues
    if "//" in v:
        issues.append("FTP Path に連続スラッシュ(//)があります")
    if " " in v:
        issues.append("FTP Path に空白が含まれています")
    if "yymmdd" not in v:
        issues.append("FTP Path に yymmdd トークンがありません")
    return issues


def resolved_remote_path(path_template: str, token: str | None = None) -> str:
    t = token or datetime.now().strftime("%y%m%d")
    return path_template.replace("yymmdd", t).strip()


def _ftp_scheme(common: CommonConfig) -> str:
    enc = common.ftp_encryption.lower()
    return "ftps" if "implicit" in enc else "ftp"


def _build_ftp_url(common: CommonConfig, remote_path: str = "") -> str:
    scheme = _ftp_scheme(common)
    path = remote_path.strip("/")
    if path:
        return f"{scheme}://{common.ftp_host}:{common.ftp_port}/{path}"
    return f"{scheme}://{common.ftp_host}:{common.ftp_port}/"


def _build_curl_base_command(common: CommonConfig, curl_path: Path, *, timeout_sec: int) -> list[str]:
    enc = common.ftp_encryption.lower()
    base = [
        str(curl_path),
        "--connect-timeout",
        str(timeout_sec),
        "--max-time",
        str(max(timeout_sec * 2, 30)),
        "--fail",
        "--show-error",
        "--disable-epsv",
        "--user",
        f"{common.ftp_username}:{common.ftp_password}",
    ]
    if "explicit" in enc:
        base.append("--ssl-reqd")
    if "implicit" in enc:
        base.extend(["--ssl-reqd", "--ftp-ssl"])
    return base


def _sanitize_command(cmd: list[str], password: str) -> str:
    rendered: list[str] = []
    for token in cmd:
        replaced = token.replace(password, "********") if password else token
        rendered.append(replaced)
    return " ".join(rendered)


def run_ftp_curl_command(common: CommonConfig, curl_path: Path, extra_args: list[str], *, timeout_sec: int = 20) -> CurlResult:
    cmd = _build_curl_base_command(common, curl_path, timeout_sec=timeout_sec)
    cmd.extend(extra_args)
    creationflags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
    p = subprocess.run(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="replace",
        creationflags=creationflags,
        check=False,
    )
    return CurlResult(
        command_summary=_sanitize_command(cmd, common.ftp_password),
        returncode=p.returncode,
        stdout=(p.stdout or "").strip(),
        stderr=(p.stderr or "").strip(),
    )


def ftp_test_connection(common: CommonConfig, curl_path: Path) -> tuple[str, CurlResult]:
    remote = resolved_remote_path(common.ftp_remote_path_template)
    test_url = _build_ftp_url(common, remote)
    result = run_ftp_curl_command(common, curl_path, ["--list-only", test_url], timeout_sec=15)
    if result.returncode != 0:
        raise RuntimeError(
            "\n".join(
                [
                    "FTP接続テスト失敗",
                    f"host={common.ftp_host}",
                    f"port={common.ftp_port}",
                    f"encryption={common.ftp_encryption}",
                    f"username={common.ftp_username}",
                    f"remote_path={remote}",
                    f"curl={result.command_summary}",
                    f"stdout={result.stdout or '(empty)'}",
                    f"stderr={result.stderr or '(empty)'}",
                ]
            )
        )
    return remote, result


def ftp_upload(local_file: Path, common: CommonConfig, curl_path: Path) -> tuple[str, CurlResult]:
    remote = resolved_remote_path(common.ftp_remote_path_template)
    remote_clean = remote.strip("/")
    target = f"{remote_clean}/{local_file.name}" if remote_clean else local_file.name
    target_url = _build_ftp_url(common, target)
    result = run_ftp_curl_command(
        common,
        curl_path,
        ["--ftp-create-dirs", "--upload-file", str(local_file), target_url],
        timeout_sec=30,
    )
    if result.returncode != 0:
        raise RuntimeError(
            "\n".join(
                [
                    "FTPアップロード失敗",
                    f"host={common.ftp_host}",
                    f"port={common.ftp_port}",
                    f"encryption={common.ftp_encryption}",
                    f"username={common.ftp_username}",
                    f"remote_path={remote}",
                    f"curl={result.command_summary}",
                    f"stdout={result.stdout or '(empty)'}",
                    f"stderr={result.stderr or '(empty)'}",
                ]
            )
        )
    return target, result


def local_copy(src: Path, common: CommonConfig) -> None:
    dst_dir = Path(common.default_local_copy_dir)
    if not str(dst_dir):
        return
    dst_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst_dir / src.name)


def sanitize_filename(value: str, max_len: int = 140) -> str:
    name = (value or "").replace("\u00a0", " ")
    for ch in '\\/:*?"<>|':
        name = name.replace(ch, "_")
    name = " ".join(name.split()).strip()
    return name[:max_len] if max_len > 0 else name


def save_inputtable_excel(
    *,
    output_path: Path,
    merged_sheet: list[list[str]],
    merged_ranges: list[dict],
    flat_sheet: list[list[str]],
) -> None:
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError("Excel出力には openpyxl が必要です。") from exc

    wb = Workbook()
    ws_table = wb.active
    ws_table.title = "Table"
    for row in merged_sheet:
        ws_table.append(list(row))
    for m in merged_ranges:
        s = m.get("s", {})
        e = m.get("e", {})
        sr = int(s.get("r", 0)) + 1
        sc = int(s.get("c", 0)) + 1
        er = int(e.get("r", 0)) + 1
        ec = int(e.get("c", 0)) + 1
        if er > sr or ec > sc:
            ws_table.merge_cells(start_row=sr, start_column=sc, end_row=er, end_column=ec)

    ws_flat = wb.create_sheet("Flat")
    for row in flat_sheet:
        ws_flat.append(list(row))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
